import os
from os.path import basename
from zipfile import ZipFile
import csv
from PyPDF2 import PdfReader
from io import TextIOWrapper
from openpyxl import load_workbook

path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files')
path_destination = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources')
path_zip = os.path.join(path_destination, "lesson_6_hw.zip")


def test_create_archive():
    file_dir = os.listdir(path)
    with ZipFile(path_zip, "w") as myzip:
        for file in file_dir:
            add_file = os.path.join(path, file)
            myzip.write(add_file, basename(add_file))


def test_read_and_content_csv():
    zf = ZipFile(path_zip)
    with zf.open("my.csv") as csvfile:
        csvfile = csv.reader(TextIOWrapper(csvfile))
        list_csv = []
        for r in csvfile:
            text = "".join(r).replace(";", " ")
            list_csv.append(text)
            assert "11|12|13|14|15|16|17|18|19|20" in list_csv, f"В файле нет этих данных {'11|12|13|14|15|16|17|18|19|20'}"
    zf.close()


def test_read_and_content_pdf():
    with ZipFile(path_zip) as zf:
        pdf_file = zf.extract("my.pdf")
        reader = PdfReader(pdf_file)
        try:
            page = reader.pages[0]
            text = page.extract_text()
            assert "Это мой PDF файл для домашнего задания №6" in text, \
                f'В файле нет строки Это мой PDF файл для домашнего задания №6'
        finally:
            os.remove(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'my.pdf'))
            zf.close()


def test_read_and_content_xlsx():
    zf = ZipFile(path_zip)
    with zf.open("my.xlsx") as xlsxfile:
        xlsxfile = load_workbook(xlsxfile)
        sheet = xlsxfile.active
        assert sheet.cell(row=2, column=3).value == 2800
    zf.close()
