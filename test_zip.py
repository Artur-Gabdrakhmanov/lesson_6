import glob
import os
from os.path import basename
from zipfile import ZipFile
import csv
import pytest
from PyPDF2 import PdfReader
from io import TextIOWrapper
from openpyxl import load_workbook


@pytest.fixture()
def clear_dir():
    precondition_directory()


path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'download')
path_destination = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources')
path_zip = os.path.join(path_destination, "artur.zip")


def test_create_archive(clear_dir):
    file_dir = os.listdir(path)
    with ZipFile(path_zip, "w") as myzip:
        for file in file_dir:
            add_file = os.path.join(path, file)
            myzip.write(add_file, basename(add_file))
    files = os.listdir(path_destination)
    assert len(files) == 1, f"Неверное количество скаченных файлов {len(files)} не ровно {1}"
    assert "artur.zip" == files[0], f"Архив {files[0]} создался с неправильным именем"


def test_read_and_content_csv():
    zf = ZipFile(path_zip)
    with zf.open("my.csv") as csvfile:
        csvfile = csv.reader(TextIOWrapper(csvfile))
        list_csv = []
        for r in csvfile:
            text = "".join(r).replace(";", " ", 3)
            list_csv.append(text)
    assert "63|57|51|45|39|33|27|21|15|9" in list_csv, f"В файле отсутсвует информация " \
                                                       f"о пользователе {'63|57|51|45|39|33|27|21|15|9'}"
    zf.close()


def test_read_and_content_pdf():
    with ZipFile(path_zip) as zf:
        pdf_file = zf.extract("my.pdf")
        reader = PdfReader(pdf_file)
        try:
            page = reader.pages[0]
            text = page.extract_text()
            #result_filters = text.split()
            #del result_filters[0]
            #for text in result_filters:
            assert "Это мой PDF файл для домашнего задания №6" in text, \
                    f'Это мой PDF файл для домашнего задания №6'
        finally:
            os.remove(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'my.pdf'))
            zf.close()


def test_read_and_content_xlsx():
    zf = ZipFile(path_zip)
    with zf.open("my.xlsx") as xlsxfile:
        xlsxfile = load_workbook(xlsxfile)
        sheet = xlsxfile.active
        assert sheet.cell(row=2, column=3).value == 2800
    zf.close()


def precondition_directory():
    path_file = os.path.join(path_destination, '*.*')
    for file in glob.glob(path_file):
        os.remove(file)